#!/usr/bin/env python3
"""
Generate enriched 14-sheet GAAIF Analysis Excel workbook.

Runs the full pricing pipeline with live/fallback market data and produces
output/GAAIF_Analysis_Data.xlsx matching the Feb 1 enriched version quality.
"""

import numpy as np
import time
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from pricing_model import MarketData, ContractTerms, StructuredForwardPricer, CorrelatedGBMSimulator
from market_data import MarketDataProvider, SensitivityRangeGenerator

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


# ── Styling helpers ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
SECTION_FONT = Font(name='Calibri', bold=True, size=11, color='1A365D')
SECTION_FILL = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='1A365D')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='E2E8F0')
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def style_section_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL


def auto_width(ws, min_width=10, max_width=25):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        width = max(min_width, min(max_width, max(lengths) if lengths else min_width))
        ws.column_dimensions[col_letter].width = width


def fmt_eur(v):
    """Format EUR value for display."""
    if abs(v) >= 1e9:
        return f"EUR {v/1e9:+,.2f}B"
    if abs(v) >= 1e6:
        return f"EUR {v/1e6:+,.2f}M"
    if abs(v) >= 1e3:
        return f"EUR {v/1e3:+,.1f}K"
    return f"EUR {v:+,.2f}"


# ── Main pipeline ────────────────────────────────────────────────────────────

def main():
    os.makedirs('output', exist_ok=True)
    out_path = 'output/GAAIF_Analysis_Data.xlsx'

    print("=" * 60)
    print("GENERATING ENRICHED 14-SHEET EXCEL WORKBOOK")
    print("=" * 60)

    # ── Fetch market data ────────────────────────────────────────────────
    print("\n[1/14] Fetching market data...")
    provider = MarketDataProvider(use_live=True)
    market = provider.fetch_market_data()
    contract = ContractTerms()
    ranges = SensitivityRangeGenerator(market, contract)
    pricer = StructuredForwardPricer(market, contract)
    print(provider.get_provenance_report())

    # ── Base pricing (100K paths) ────────────────────────────────────────
    print("\n[2/14] Running base Monte Carlo (100K paths)...")
    t0 = time.time()
    base = pricer.price_monte_carlo(n_paths=100000, seed=42)
    base_time = time.time() - t0
    print(f"  PV = {fmt_eur(base['price_zgroup'])}, KO = {base['knockout_rate']*100:.1f}%, time = {base_time:.1f}s")

    # ── Greeks ───────────────────────────────────────────────────────────
    print("\n[3/14] Computing Greeks (50K paths per bump)...")
    greeks = pricer.compute_greeks(base_result=base, n_paths=50000, seed=42)

    # ── Create workbook ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # =====================================================================
    # Sheet 1: Executive Summary
    # =====================================================================
    print("\n[4/14] Writing Executive Summary...")
    ws = wb.active
    ws.title = "Executive Summary"

    rows = [
        ("GAAIF Challenge 2026 - Structured Gold Forward Pricing Analysis", None),
        (None, None),
        ("ANALYSIS OVERVIEW", None),
        ("Analysis Date", f"February 26, 2026"),
        ("Prepared By", "GAAIF Challenge Submission"),
        ("Model", "Two-Factor Correlated GBM with Knock-Out Barriers"),
        ("Simulation", f"100,000 paths, 504 time steps"),
        (None, None),
        ("PRODUCT SUMMARY", None),
        ("Product Type", "Structured Gold Forward with Double Knock-Out"),
        ("Notional Principal", "EUR 500,000,000"),
        ("Strike Price", f"USD {contract.strike:,.2f}/oz"),
        ("Tenor", f"{contract.tenor:.0f} years"),
        ("Lower Barrier (EUR/USD)", f"{contract.barrier_lower}"),
        ("Upper Barrier (EUR/USD)", f"{contract.barrier_upper}"),
        (None, None),
        ("CURRENT MARKET", None),
        ("Gold Spot", f"${market.gold_spot:,.2f}/oz"),
        ("EUR/USD Spot", f"{market.eurusd_spot:.4f}"),
        ("Gold Volatility", f"{market.sigma_gold*100:.1f}%"),
        ("EUR/USD Volatility", f"{market.sigma_eurusd*100:.1f}%"),
        (None, None),
        ("VALUATION RESULTS", None),
        ("Z Group Present Value", fmt_eur(base['price_zgroup'])),
        ("A Bank Present Value", fmt_eur(base['price_abank'])),
        ("Standard Error", fmt_eur(base['std_error'])),
        ("95% CI Lower", fmt_eur(base['ci_95_lower'])),
        ("95% CI Upper", fmt_eur(base['ci_95_upper'])),
        (None, None),
        ("KNOCKOUT ANALYSIS", None),
        ("Knockout Probability", f"{base['knockout_rate']*100:.2f}%"),
        ("Average KO Time", f"{base['avg_knockout_time']:.3f} years"),
        ("Upper Barrier Breaches", f"{base['upper_breach_rate']*100:.1f}%"),
        ("Lower Barrier Breaches", f"{base['lower_breach_rate']*100:.1f}%"),
        ("No Knockout Rate", f"{(1-base['knockout_rate'])*100:.2f}%"),
    ]

    for r, (a, b) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)

    # Style title
    ws.cell(row=1, column=1).font = TITLE_FONT
    for section_label in ["ANALYSIS OVERVIEW", "PRODUCT SUMMARY", "CURRENT MARKET",
                          "VALUATION RESULTS", "KNOCKOUT ANALYSIS"]:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
            for cell in row:
                if cell.value == section_label:
                    style_section_row(ws, cell.row, 2)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35

    # =====================================================================
    # Sheet 2: Market Data
    # =====================================================================
    print("[5/14] Writing Market Data...")
    ws = wb.create_sheet("Market Data")

    headers = ['Parameter', 'Symbol', 'Value', 'Units', 'Source', 'Notes']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 6)

    prov = provider._provenance
    gold_prov = prov.get('gold_spot')
    fx_prov = prov.get('eurusd_spot')

    market_rows = [
        ('SPOT PRICES', None, None, None, None, None),
        ('Gold Spot Price', 'S₀', market.gold_spot, 'USD/oz',
         f'{gold_prov.source.upper()} ({gold_prov.ticker})' if gold_prov else 'Fallback',
         'Current spot price'),
        ('EUR/USD Spot Rate', 'X₀', market.eurusd_spot, None,
         f'{fx_prov.source.upper()} ({fx_prov.ticker})' if fx_prov else 'Fallback',
         f'Distance to upper barrier: {(contract.barrier_upper - market.eurusd_spot)/market.eurusd_spot*100:.1f}%'),
        (None, None, None, None, None, None),
        ('INTEREST RATES', None, None, None, None, None),
        ('EUR Risk-Free Rate', 'r_EUR', market.r_eur, '%/year',
         f'{prov["r_eur"].source.upper()}', 'ECB deposit rate'),
        ('USD Risk-Free Rate', 'r_USD', market.r_usd, '%/year',
         f'{prov["r_usd"].source.upper()} ({prov["r_usd"].ticker})', '13-week T-bill proxy'),
        ('Rate Differential', 'r_EUR - r_USD',
         market.r_eur - market.r_usd, '%/year', 'Calculated', 'Drives FX drift'),
        (None, None, None, None, None, None),
        ('VOLATILITIES', None, None, None, None, None),
        ('Gold Volatility', 'σ_S', market.sigma_gold, 'annualized',
         f'{prov["sigma_gold"].source.upper()}', 'EWMA (λ=0.94)'),
        ('EUR/USD Volatility', 'σ_X', market.sigma_eurusd, 'annualized',
         f'{prov["sigma_eurusd"].source.upper()}', 'EWMA (λ=0.94)'),
        (None, None, None, None, None, None),
        ('CORRELATION & YIELD', None, None, None, None, None),
        ('Gold-EURUSD Correlation', 'ρ', market.rho, None,
         f'{prov["rho"].source.upper()}', f'{prov["rho"].ticker}'),
        ('Gold Convenience Yield', 'q', market.gold_yield, 'annualized',
         f'{prov["gold_yield"].source.upper()}', 'From futures term structure'),
        (None, None, None, None, None, None),
        ('DERIVED PARAMETERS', None, None, None, None, None),
        ('Gold Forward (2Y)', 'F(0,T)', market.gold_spot * np.exp((market.r_usd - market.gold_yield) * contract.tenor),
         'USD/oz', 'Calculated', 'S₀ × exp((r_USD - q) × T)'),
        ('Moneyness', 'S₀/K', market.gold_spot / contract.strike, None,
         'Calculated', f'Gold {(market.gold_spot/contract.strike - 1)*100:+.1f}% vs strike'),
        ('Intrinsic Value', 'N×(S-K)/K',
         contract.notional * (market.gold_spot - contract.strike) / contract.strike,
         'EUR', 'Calculated', 'If settled today'),
    ]

    for r, row_data in enumerate(market_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        if row_data[0] and row_data[1] is None and row_data[2] is None:
            style_section_row(ws, r, 6)

    auto_width(ws)

    # =====================================================================
    # Sheet 3: Contract Terms
    # =====================================================================
    print("[6/14] Writing Contract Terms...")
    ws = wb.create_sheet("Contract Terms")

    headers = ['Parameter', 'Value', 'Description']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 3)

    strike_vs_spot = (contract.strike - market.gold_spot) / market.gold_spot * 100
    dist_to_lower = (market.eurusd_spot - contract.barrier_lower) / market.eurusd_spot * 100
    dist_to_upper = (contract.barrier_upper - market.eurusd_spot) / market.eurusd_spot * 100

    terms_rows = [
        ('PRINCIPAL & STRIKE', None, None),
        ('Notional Principal', 'EUR 500,000,000', 'EUR 500 Million'),
        ('Strike Price (K)', f'USD {contract.strike:,.2f}/oz', 'Fixed gold reference price'),
        ('Strike vs Current Spot', f'{strike_vs_spot:+.2f}%', 'Strike relative to current gold spot'),
        (None, None, None),
        ('TENOR', None, None),
        ('Contract Tenor', '2.0 years', 'March 2026 - February 2028'),
        ('Time Steps', '504', '252 trading days × 2 years'),
        (None, None, None),
        ('BARRIERS (EUR/USD)', None, None),
        ('Lower Barrier', f'{contract.barrier_lower}', 'Knock-out if EUR/USD < 1.05'),
        ('Upper Barrier', f'{contract.barrier_upper}', 'Knock-out if EUR/USD > 1.25'),
        ('Barrier Width', f'{contract.barrier_upper - contract.barrier_lower:.2f}', 'Upper - Lower'),
        ('Current Distance to Lower', f'{dist_to_lower:.1f}%', f'EUR/USD {market.eurusd_spot:.4f} vs {contract.barrier_lower}'),
        ('Current Distance to Upper', f'{dist_to_upper:.1f}%', f'EUR/USD {market.eurusd_spot:.4f} vs {contract.barrier_upper}'),
        (None, None, None),
        ('SETTLEMENT', None, None),
        ('Z Group Payoff', 'N × (P - K) / K', 'Positive when gold > strike'),
        ('A Bank Payoff', 'N × (K - P) / K', 'Positive when gold < strike'),
        ('Settlement Price (P)', 'Gold at KO or maturity', 'Path-dependent'),
        (None, None, None),
        ('MONITORING', None, None),
        ('Barrier Monitoring', 'Continuous', 'Checked at each simulation step'),
        ('Settlement on KO', 'Immediate', 'Gold price at barrier breach time'),
    ]

    for r, row_data in enumerate(terms_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        if row_data[0] and row_data[1] is None:
            style_section_row(ws, r, 3)

    auto_width(ws, max_width=40)

    # =====================================================================
    # Sheet 4: Monte Carlo Results
    # =====================================================================
    print("[7/14] Writing Monte Carlo Results...")
    ws = wb.create_sheet("Monte Carlo Results")

    headers = ['Metric', 'Value', 'Units', 'Notes']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 4)

    mc_rows = [
        ('VALUATION', None, None, None),
        ('Z Group Present Value', base['price_zgroup'], 'EUR', 'Positive = asset for Z Group'),
        ('A Bank Present Value', base['price_abank'], 'EUR', 'Negative = liability for A Bank'),
        ('Absolute Value', abs(base['price_zgroup']), 'EUR', 'Position size'),
        (None, None, None, None),
        ('STATISTICAL CONFIDENCE', None, None, None),
        ('Standard Error', base['std_error'], 'EUR', f'SE = σ/√n, n={base["n_paths"]:,}'),
        ('Relative SE', base['std_error'] / abs(base['price_zgroup']) * 100 if base['price_zgroup'] != 0 else 0, '%', 'SE / |PV|'),
        ('95% CI Lower', base['ci_95_lower'], 'EUR', 'PV - 1.96 × SE'),
        ('95% CI Upper', base['ci_95_upper'], 'EUR', 'PV + 1.96 × SE'),
        ('CI Width', base['ci_95_upper'] - base['ci_95_lower'], 'EUR', '95% confidence band'),
        (None, None, None, None),
        ('KNOCKOUT STATISTICS', None, None, None),
        ('Knockout Rate', base['knockout_rate'], None, f'{base["knockout_rate"]*100:.2f}% of paths'),
        ('No-Knockout Rate', 1 - base['knockout_rate'], None, 'Paths reaching maturity'),
        ('Average KO Time', base['avg_knockout_time'], 'years', 'Mean time to barrier breach'),
        ('Upper Breach Rate', base['upper_breach_rate'], None, 'EUR/USD > 1.25 breaches'),
        ('Lower Breach Rate', base['lower_breach_rate'], None, 'EUR/USD < 1.05 breaches'),
        (None, None, None, None),
        ('SETTLEMENT STATISTICS', None, None, None),
        ('Mean Settlement Price', np.mean(base['settlement_prices']), 'USD/oz', 'Average gold at settlement'),
        ('Std Settlement Price', np.std(base['settlement_prices']), 'USD/oz', 'Settlement price dispersion'),
        ('Mean Settlement Time', np.mean(base['settlement_times']), 'years', 'Average time to settlement'),
        (None, None, None, None),
        ('SIMULATION PARAMETERS', None, None, None),
        ('Number of Paths', base['n_paths'], None, 'Monte Carlo paths'),
        ('Time Steps', base['n_steps'], None, 'Per path'),
    ]

    for r, row_data in enumerate(mc_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        if row_data[0] and row_data[1] is None and row_data[2] is None:
            style_section_row(ws, r, 4)

    auto_width(ws)

    # =====================================================================
    # Sheet 5: Greeks
    # =====================================================================
    print("[8/14] Writing Greeks...")
    ws = wb.create_sheet("Greeks")

    headers = ['Greek', 'Value', 'Per Unit', 'Scaled Impact', 'Interpretation']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 5)

    greeks_rows = [
        ('Delta (Gold)', greeks['delta_gold'], 'EUR per $1',
         greeks['delta_gold'] * 100, '$100 move impact (EUR)'),
        ('Gamma (Gold)', greeks['gamma_gold'], 'EUR per $1²',
         None, 'Convexity (second derivative)'),
        ('Delta (EUR/USD)', greeks['delta_eurusd'], 'EUR per 0.01 FX',
         greeks['delta_eurusd'] * (contract.barrier_upper - market.eurusd_spot),
         f'Move to {contract.barrier_upper} barrier impact'),
        ('Vega (Gold)', greeks['vega_gold'], 'EUR per 1% vol',
         greeks['vega_gold'] * 5, '5% vol increase impact'),
        ('Rho (EUR)', greeks['rho_eur'], 'EUR per 1% rate',
         greeks['rho_eur'] * 0.001, '10bp rate change impact'),
        ('Rho (USD)', greeks.get('rho_usd', 0), 'EUR per 1% rate',
         greeks.get('rho_usd', 0) * 0.001, '10bp rate change impact'),
        ('Correlation Sensitivity', greeks.get('correlation_sensitivity', 0),
         'EUR per 0.01 corr', greeks.get('correlation_sensitivity', 0) * 5,
         '0.05 correlation shift impact'),
    ]

    for r, row_data in enumerate(greeks_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    auto_width(ws)

    # =====================================================================
    # Sheet 6: Gold Sensitivity (20 points)
    # =====================================================================
    print("[9/14] Running Gold Sensitivity (20 points)...")
    ws = wb.create_sheet("Gold Sensitivity")

    headers = ['Gold Price (USD/oz)', 'Z Group PV (EUR)', 'A Bank PV (EUR)',
               'Knockout Rate', 'Intrinsic Value (EUR)', 'Time Value (EUR)']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 6)

    gold_spots = ranges.gold_spot_range(20)
    for i, gs in enumerate(gold_spots):
        m = MarketData(gold_spot=gs, eurusd_spot=market.eurusd_spot,
                       r_eur=market.r_eur, r_usd=market.r_usd,
                       sigma_gold=market.sigma_gold, sigma_eurusd=market.sigma_eurusd,
                       rho=market.rho, gold_yield=market.gold_yield)
        p = StructuredForwardPricer(m, contract)
        res = p.price_monte_carlo(n_paths=30000, seed=42)
        intrinsic = contract.notional * (gs - contract.strike) / contract.strike
        r_idx = i + 2
        ws.cell(row=r_idx, column=1, value=gs)
        ws.cell(row=r_idx, column=2, value=res['price_zgroup'])
        ws.cell(row=r_idx, column=3, value=res['price_abank'])
        ws.cell(row=r_idx, column=4, value=res['knockout_rate'])
        ws.cell(row=r_idx, column=5, value=intrinsic)
        ws.cell(row=r_idx, column=6, value=res['price_zgroup'] - intrinsic)
        if (i + 1) % 5 == 0:
            print(f"   Gold ${gs:.0f}: PV = {fmt_eur(res['price_zgroup'])}")

    auto_width(ws)

    # =====================================================================
    # Sheet 7: EURUSD Sensitivity (20 points)
    # =====================================================================
    print("[10/14] Running EUR/USD Sensitivity (20 points)...")
    ws = wb.create_sheet("EURUSD Sensitivity")

    headers = ['EUR/USD Rate', 'Z Group PV (EUR)', 'Knockout Rate',
               'Upper Breach %', 'Lower Breach %', 'Distance to Upper', 'Distance to Lower']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 7)

    fx_spots = ranges.eurusd_spot_range(20)
    for i, fx in enumerate(fx_spots):
        m = MarketData(gold_spot=market.gold_spot, eurusd_spot=fx,
                       r_eur=market.r_eur, r_usd=market.r_usd,
                       sigma_gold=market.sigma_gold, sigma_eurusd=market.sigma_eurusd,
                       rho=market.rho, gold_yield=market.gold_yield)
        p = StructuredForwardPricer(m, contract)
        res = p.price_monte_carlo(n_paths=30000, seed=42)
        r_idx = i + 2
        ws.cell(row=r_idx, column=1, value=fx)
        ws.cell(row=r_idx, column=2, value=res['price_zgroup'])
        ws.cell(row=r_idx, column=3, value=res['knockout_rate'])
        ws.cell(row=r_idx, column=4, value=res['upper_breach_rate'])
        ws.cell(row=r_idx, column=5, value=res['lower_breach_rate'])
        ws.cell(row=r_idx, column=6, value=f"{(contract.barrier_upper - fx)/fx*100:.1f}%")
        ws.cell(row=r_idx, column=7, value=f"{(fx - contract.barrier_lower)/fx*100:.1f}%")
        if (i + 1) % 5 == 0:
            print(f"   EUR/USD {fx:.3f}: PV = {fmt_eur(res['price_zgroup'])}")

    auto_width(ws)

    # =====================================================================
    # Sheet 8: Volatility Grid (6 gold_vol × 5 fx_vol)
    # =====================================================================
    print("[11/14] Running Volatility Grid (6×5)...")
    ws = wb.create_sheet("Volatility Grid")

    gold_vols = ranges.gold_vol_range(6)
    fx_vols = ranges.eurusd_vol_range(5)

    # Header row: "Gold Vol \ FX Vol", then FX vol labels
    ws.cell(row=1, column=1, value='Gold Vol \\ FX Vol')
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    for j, fv in enumerate(fx_vols):
        ws.cell(row=1, column=j+2, value=f'{fv*100:.0f}%')
        ws.cell(row=1, column=j+2).font = HEADER_FONT
        ws.cell(row=1, column=j+2).fill = HEADER_FILL

    for i, gv in enumerate(gold_vols):
        ws.cell(row=i+2, column=1, value=f'{gv*100:.0f}%')
        ws.cell(row=i+2, column=1).font = Font(bold=True)
        for j, fv in enumerate(fx_vols):
            m = MarketData(gold_spot=market.gold_spot, eurusd_spot=market.eurusd_spot,
                           r_eur=market.r_eur, r_usd=market.r_usd,
                           sigma_gold=gv, sigma_eurusd=fv,
                           rho=market.rho, gold_yield=market.gold_yield)
            p = StructuredForwardPricer(m, contract)
            res = p.price_monte_carlo(n_paths=20000, seed=42)
            ws.cell(row=i+2, column=j+2, value=res['price_zgroup'])
        print(f"   Gold Vol {gv*100:.0f}% row complete")

    auto_width(ws)

    # =====================================================================
    # Sheet 9: Correlation Sensitivity (15 points)
    # =====================================================================
    print("[12/14] Running Correlation Sensitivity (15 points)...")
    ws = wb.create_sheet("Correlation Sensitivity")

    headers = ['Correlation (ρ)', 'Z Group PV (EUR)', 'Knockout Rate',
               'Upper Breach %', 'Lower Breach %']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 5)

    correlations = ranges.correlation_range(15)
    for i, rho in enumerate(correlations):
        m = MarketData(gold_spot=market.gold_spot, eurusd_spot=market.eurusd_spot,
                       r_eur=market.r_eur, r_usd=market.r_usd,
                       sigma_gold=market.sigma_gold, sigma_eurusd=market.sigma_eurusd,
                       rho=rho, gold_yield=market.gold_yield)
        p = StructuredForwardPricer(m, contract)
        res = p.price_monte_carlo(n_paths=30000, seed=42)
        r_idx = i + 2
        ws.cell(row=r_idx, column=1, value=rho)
        ws.cell(row=r_idx, column=2, value=res['price_zgroup'])
        ws.cell(row=r_idx, column=3, value=res['knockout_rate'])
        ws.cell(row=r_idx, column=4, value=res['upper_breach_rate'])
        ws.cell(row=r_idx, column=5, value=res['lower_breach_rate'])

    auto_width(ws)

    # =====================================================================
    # Sheet 10: Convergence Analysis
    # =====================================================================
    print("[13/14] Running Convergence Analysis...")
    ws = wb.create_sheet("Convergence Analysis")

    headers = ['Number of Paths', 'Z Group PV (EUR)', 'Standard Error (EUR)',
               'Relative SE (%)', '95% CI Width (EUR)']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 5)

    path_counts = [1000, 2500, 5000, 10000, 20000, 50000, 100000, 150000, 200000]
    for i, n in enumerate(path_counts):
        t0 = time.time()
        res = pricer.price_monte_carlo(n_paths=n, seed=42)
        elapsed = time.time() - t0
        rel_se = res['std_error'] / abs(res['price_zgroup']) * 100 if res['price_zgroup'] != 0 else 0
        ci_width = res['ci_95_upper'] - res['ci_95_lower']
        r_idx = i + 2
        ws.cell(row=r_idx, column=1, value=n)
        ws.cell(row=r_idx, column=2, value=res['price_zgroup'])
        ws.cell(row=r_idx, column=3, value=res['std_error'])
        ws.cell(row=r_idx, column=4, value=rel_se)
        ws.cell(row=r_idx, column=5, value=ci_width)
        print(f"   {n:>7,} paths: PV = {fmt_eur(res['price_zgroup'])}, SE = {fmt_eur(res['std_error'])}, time = {elapsed:.1f}s")

    auto_width(ws)

    # =====================================================================
    # Sheet 11: Scenario Analysis (18 scenarios)
    # =====================================================================
    print("[14/14] Running Scenario Analysis (18 scenarios)...")
    ws = wb.create_sheet("Scenario Analysis")

    headers = ['Scenario', 'Gold (USD/oz)', 'EUR/USD', 'Gold Vol', 'FX Vol',
               'Correlation', 'Z Group PV (EUR)', 'A Bank PV (EUR)', 'KO Rate', 'Avg KO Time']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 10)

    scenarios = [
        ("Base Case - Current Market", market.gold_spot, market.eurusd_spot,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("Gold at Strike ($4,600)", contract.strike, market.eurusd_spot,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("Gold Mild Decline ($4,400)", 4400, market.eurusd_spot,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("Gold Crash ($4,000)", 4000, market.eurusd_spot,
         0.35, market.sigma_eurusd, market.rho),
        ("Gold Rally +10%", market.gold_spot * 1.10, market.eurusd_spot,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("Gold Rally +20%", market.gold_spot * 1.20, market.eurusd_spot,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("Gold Crash -20%", market.gold_spot * 0.80, market.eurusd_spot,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("EUR Strengthens to 1.22", market.gold_spot, 1.22,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("EUR Near Upper Barrier", market.gold_spot, 1.24,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("EUR Near Lower Barrier", market.gold_spot, 1.06,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("EUR Weakens to 1.10", market.gold_spot, 1.10,
         market.sigma_gold, market.sigma_eurusd, market.rho),
        ("High Vol Environment", market.gold_spot, market.eurusd_spot,
         0.45, 0.15, market.rho),
        ("Low Vol Environment", market.gold_spot, market.eurusd_spot,
         0.20, 0.06, market.rho),
        ("Combined Stress (Bull)", market.gold_spot * 1.15, 1.22,
         0.32, market.sigma_eurusd, market.rho),
        ("Combined Stress (Bear)", market.gold_spot * 0.85, 1.08,
         0.25, market.sigma_eurusd, market.rho),
        ("Tail Risk: All Adverse", market.gold_spot * 0.75, 1.06,
         0.40, 0.14, market.rho),
        ("Best Case Scenario", market.gold_spot * 1.25, 1.15,
         0.22, 0.08, market.rho),
        ("Worst Case Scenario", market.gold_spot * 0.70, 1.07,
         0.35, 0.12, market.rho),
    ]

    for i, (name, gs, fx, gv, fv, rho) in enumerate(scenarios):
        m = MarketData(gold_spot=gs, eurusd_spot=fx,
                       r_eur=market.r_eur, r_usd=market.r_usd,
                       sigma_gold=gv, sigma_eurusd=fv,
                       rho=rho, gold_yield=market.gold_yield)
        p = StructuredForwardPricer(m, contract)
        res = p.price_monte_carlo(n_paths=25000, seed=42)
        r_idx = i + 2
        ws.cell(row=r_idx, column=1, value=name)
        ws.cell(row=r_idx, column=2, value=gs)
        ws.cell(row=r_idx, column=3, value=fx)
        ws.cell(row=r_idx, column=4, value=gv)
        ws.cell(row=r_idx, column=5, value=fv)
        ws.cell(row=r_idx, column=6, value=rho)
        ws.cell(row=r_idx, column=7, value=res['price_zgroup'])
        ws.cell(row=r_idx, column=8, value=res['price_abank'])
        ws.cell(row=r_idx, column=9, value=res['knockout_rate'])
        ws.cell(row=r_idx, column=10, value=res['avg_knockout_time'])
        print(f"   {name}: PV = {fmt_eur(res['price_zgroup'])}")

    auto_width(ws, max_width=30)

    # =====================================================================
    # Sheet 12: Sample Paths (10 gold + 10 EURUSD × 21 time steps)
    # =====================================================================
    print("\nGenerating Sample Paths...")
    ws = wb.create_sheet("Sample Paths")

    n_sample = 10
    n_time_steps = 20  # 21 points including t=0
    simulator = CorrelatedGBMSimulator(market, contract, seed=42)
    sim = simulator.simulate_paths(n_paths=1000, n_steps=n_time_steps * 25)

    # Downsample to 21 time points
    step_indices = np.linspace(0, sim['gold_paths'].shape[1] - 1, n_time_steps + 1, dtype=int)
    time_points = sim['times'][step_indices]

    # Headers
    headers = ['Time (Years)']
    for i in range(1, n_sample + 1):
        headers.append(f'Gold Path {i}')
    for i in range(1, n_sample + 1):
        headers.append(f'EURUSD Path {i}')

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    for t_idx, si in enumerate(step_indices):
        r_idx = t_idx + 2
        ws.cell(row=r_idx, column=1, value=round(time_points[t_idx], 2))
        for p in range(n_sample):
            ws.cell(row=r_idx, column=p + 2, value=sim['gold_paths'][p, si])
            ws.cell(row=r_idx, column=n_sample + p + 2, value=sim['eurusd_paths'][p, si])

    auto_width(ws, min_width=12)

    # =====================================================================
    # Sheet 13: Valuation History
    # =====================================================================
    print("Writing Valuation History...")
    ws = wb.create_sheet("Valuation History")

    headers = ['Analysis Date', 'Gold Spot', 'EUR/USD', 'Gold Vol',
               'Z Group PV (EUR)', 'A Bank PV (EUR)', 'KO Rate', 'Notes']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 8)

    history_rows = [
        ('January 2026 (Original)', 2750, 1.08, 0.18,
         -191903454, 191903454, 0.926, 'Original analysis'),
        ('February 1, 2026', 4900, 1.19, 0.28,
         46330626, -46330626, 0.948, 'After gold rally'),
        (f'February 26, 2026 (Current)', market.gold_spot, market.eurusd_spot, market.sigma_gold,
         base['price_zgroup'], base['price_abank'], base['knockout_rate'],
         'Current live/fallback data'),
        (None, None, None, None, None, None, None, None),
        ('CHANGE ANALYSIS', None, None, None, None, None, None, None),
        ('Original → Feb 1',
         4900 - 2750, 1.19 - 1.08, 0.28 - 0.18,
         46330626 - (-191903454), None, None, 'Gold +78%, EUR/USD +10%'),
        ('Feb 1 → Feb 26',
         market.gold_spot - 4900, market.eurusd_spot - 1.19, market.sigma_gold - 0.28,
         base['price_zgroup'] - 46330626, None, None,
         f'Gold {(market.gold_spot-4900)/4900*100:+.1f}%'),
        ('Original → Feb 26',
         market.gold_spot - 2750, market.eurusd_spot - 1.08, market.sigma_gold - 0.18,
         base['price_zgroup'] - (-191903454), None, None,
         'Total change since original'),
    ]

    for r, row_data in enumerate(history_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        if row_data[0] == 'CHANGE ANALYSIS':
            style_section_row(ws, r, 8)

    auto_width(ws)

    # =====================================================================
    # Sheet 14: Data Provenance
    # =====================================================================
    print("Writing Data Provenance...")
    ws = wb.create_sheet("Data Provenance")

    headers = ['Parameter', 'Source', 'Ticker', 'Raw Value', 'Timestamp', 'Status']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, 6)

    param_labels = {
        'gold_spot': 'Gold Spot (USD/oz)',
        'eurusd_spot': 'EUR/USD Spot',
        'r_usd': 'USD Risk-Free Rate',
        'r_eur': 'EUR Risk-Free Rate',
        'sigma_gold': 'Gold Volatility (EWMA)',
        'sigma_eurusd': 'EUR/USD Volatility (EWMA)',
        'rho': 'Gold-EURUSD Correlation',
        'gold_yield': 'Gold Convenience Yield',
    }

    for i, (key, label) in enumerate(param_labels.items()):
        p = provider._provenance.get(key)
        if p:
            r_idx = i + 2
            ws.cell(row=r_idx, column=1, value=label)
            ws.cell(row=r_idx, column=2, value=p.source.upper())
            ws.cell(row=r_idx, column=3, value=p.ticker)
            ws.cell(row=r_idx, column=4, value=p.raw_value)
            ws.cell(row=r_idx, column=5, value=p.timestamp)
            ws.cell(row=r_idx, column=6, value='OK' if p.source == 'live' else 'FALLBACK')

    auto_width(ws)

    # ── Save workbook ────────────────────────────────────────────────────
    wb.save(out_path)
    print(f"\n{'='*60}")
    print(f"WORKBOOK SAVED: {out_path}")
    print(f"Sheets: {len(wb.sheetnames)}")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i:2d}. {name}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
